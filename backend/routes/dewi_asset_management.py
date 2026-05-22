"""
CV. Dewi Aditya ERP — Asset Management Portal

Manajemen aset perusahaan lengkap dengan:
- Registrasi & kategorisasi aset
- Penugasan ke karyawan/departemen
- Riwayat pemeliharaan (maintenance)
- Integrasi Finance (journal pembelian + depresiasi otomatis)
- Dashboard ringkas

Collections:
  dewi_assets           — master aset
  dewi_asset_categories — kategori aset
  dewi_asset_assignments — riwayat penugasan
  dewi_asset_maintenance — riwayat pemeliharaan
  dewi_asset_depreciation — riwayat posting depresiasi per periode

Endpoints:
  GET    /api/assets/dashboard         — statistik aset
  GET    /api/assets/categories        — list kategori
  POST   /api/assets/categories        — create kategori
  PUT    /api/assets/categories/{id}   — update kategori
  DELETE /api/assets/categories/{id}   — delete kategori
  GET    /api/assets                   — list aset (paginated + filter)
  POST   /api/assets                   — daftarkan aset baru (+ auto journal)
  GET    /api/assets/{id}              — detail aset + jadwal depresiasi
  PUT    /api/assets/{id}              — update aset
  POST   /api/assets/{id}/assign       — tugaskan aset ke karyawan
  POST   /api/assets/{id}/unassign     — kembalikan aset
  GET    /api/assets/{id}/assignments  — riwayat penugasan
  POST   /api/assets/{id}/maintenance  — catat pemeliharaan
  GET    /api/assets/{id}/maintenance  — riwayat pemeliharaan
  POST   /api/assets/{id}/depreciate/{period} — posting depresiasi (YYYY-MM)
  GET    /api/assets/{id}/depreciation-history — riwayat depresiasi
  POST   /api/assets/{id}/dispose      — dispose aset
  GET    /api/assets/my-assets         — aset yang ditugaskan ke saya
"""
from fastapi import APIRouter, Request, HTTPException, Query, UploadFile, File
from database import get_db
from auth import require_auth, serialize_doc
from storage import put_object, generate_storage_path
from datetime import datetime, timezone, date
from dateutil.relativedelta import relativedelta
from typing import Optional
from routes.shared import paginated_response
import uuid
import math
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/assets", tags=["asset-management"])


def _uid(): return str(uuid.uuid4())
def _now(): return datetime.now(timezone.utc)


def _ser(doc):
    if not doc:
        return doc
    doc = {k: v for k, v in doc.items() if k != '_id'}
    for k, v in doc.items():
        if isinstance(v, datetime):
            doc[k] = v.isoformat()
    return doc


DEFAULT_CATEGORIES = [
    {"name": "Peralatan IT",    "code": "IT",  "useful_life_years": 4,  "depr_method": "straight_line"},
    {"name": "Mesin Produksi",  "code": "MP",  "useful_life_years": 10, "depr_method": "straight_line"},
    {"name": "Kendaraan",       "code": "KD",  "useful_life_years": 5,  "depr_method": "double_declining"},
    {"name": "Bangunan",        "code": "BG",  "useful_life_years": 20, "depr_method": "straight_line"},
    {"name": "Perabot & Mebel", "code": "PM",  "useful_life_years": 8,  "depr_method": "straight_line"},
    {"name": "Alat & Perkakas", "code": "AP",  "useful_life_years": 5,  "depr_method": "straight_line"},
    {"name": "Lain-lain",       "code": "LN",  "useful_life_years": 5,  "depr_method": "straight_line"},
]


async def _ensure_default_categories(db):
    cnt = await db.dewi_asset_categories.count_documents({})
    if cnt == 0:
        for cat in DEFAULT_CATEGORIES:
            cat["id"] = _uid()
            cat["created_at"] = _now()
        await db.dewi_asset_categories.insert_many(DEFAULT_CATEGORIES)


async def _gen_asset_number(db, category_code: str) -> str:
    year = date.today().year
    prefix = f"AST-{category_code}-{year}-"
    cnt = await db.dewi_assets.count_documents({"asset_number": {"$regex": f"^{prefix}"}})
    return f"{prefix}{str(cnt + 1).zfill(4)}"


def _calc_straight_line_monthly(cost: float, residual: float, life_months: int) -> float:
    if life_months <= 0:
        return 0.0
    return round((cost - residual) / life_months, 2)


def _calc_nbv(asset: dict) -> float:
    """Calculate Net Book Value = purchase_cost - accumulated_depreciation."""
    return round(
        float(asset.get("purchase_cost", 0)) - float(asset.get("accumulated_depreciation", 0)),
        2
    )


async def _create_finance_journal(db, user_id: str, user_name: str, date_str: str,
                                   memo: str, lines: list, source_module: str = "asset_management",
                                   source_ref: str = None) -> Optional[str]:
    """Create a journal entry in rahaza_journal_entries. Returns je_id."""
    try:
        # Generate JE number
        year_prefix = date_str[:7].replace("-", "")
        cnt = await db.rahaza_journal_entries.count_documents(
            {"je_number": {"$regex": f"^JE-{year_prefix}"}}
        )
        je_number = f"JE-{year_prefix}-{str(cnt + 1).zfill(5)}"
        je_id = _uid()
        total_debit = sum(float(ln.get("debit", 0)) for ln in lines)
        total_credit = sum(float(ln.get("credit", 0)) for ln in lines)
        doc = {
            "id": je_id,
            "je_number": je_number,
            "date": date_str,
            "memo": memo,
            "source_module": source_module,
            "source_ref": source_ref,
            "status": "draft",
            "total_debit": total_debit,
            "total_credit": total_credit,
            "created_at": _now(),
            "updated_at": _now(),
            "created_by": user_id,
            "created_by_name": user_name,
            "posted_at": None,
            "posted_by": None,
            "voided_at": None,
            "voided_by": None,
            "lines": [
                {
                    "line_id": _uid(),
                    "account_code": ln.get("account_code", ""),
                    "account_name": ln.get("account_name", ""),
                    "account_type": ln.get("account_type", "asset"),
                    "debit": float(ln.get("debit", 0)),
                    "credit": float(ln.get("credit", 0)),
                    "description": ln.get("description", ""),
                    "cost_center_id": None,
                } for ln in lines
            ]
        }
        await db.rahaza_journal_entries.insert_one(doc)
        return je_id
    except Exception as e:
        logger.warning(f"[AssetMgmt] Journal creation failed: {e}")
        return None


# ─── Dashboard ────────────────────────────────────────────────────────────

@router.get("/dashboard")
async def get_asset_dashboard(request: Request):
    user = await require_auth(request)
    db = get_db()
    await _ensure_default_categories(db)

    total = await db.dewi_assets.count_documents({"status": {"$ne": "disposed"}})
    active = await db.dewi_assets.count_documents({"status": "active"})
    in_maintenance = await db.dewi_assets.count_documents({"status": "in_maintenance"})
    disposed = await db.dewi_assets.count_documents({"status": "disposed"})

    # Total purchase cost & NBV
    pipeline = [
        {"$match": {"status": {"$ne": "disposed"}}},
        {"$group": {
            "_id": None,
            "total_cost": {"$sum": "$purchase_cost"},
            "total_accumulated_depr": {"$sum": "$accumulated_depreciation"}
        }}
    ]
    agg = await db.dewi_assets.aggregate(pipeline).to_list(1)
    totals = agg[0] if agg else {"total_cost": 0, "total_accumulated_depr": 0}
    nbv = totals["total_cost"] - totals["total_accumulated_depr"]

    # By category
    cat_pipeline = [
        {"$match": {"status": {"$ne": "disposed"}}},
        {"$group": {"_id": "$category_name", "count": {"$sum": 1}, "total_cost": {"$sum": "$purchase_cost"}}},
        {"$sort": {"count": -1}}
    ]
    by_category = await db.dewi_assets.aggregate(cat_pipeline).to_list(20)

    # Recent (5 terbaru)
    recent = await db.dewi_assets.find(
        {}, {"_id": 0, "id": 1, "name": 1, "asset_number": 1, "status": 1, "purchase_cost": 1, "category_name": 1, "created_at": 1}
    ).sort("created_at", -1).limit(5).to_list(5)

    # Depresiasi bulan ini
    current_period = date.today().strftime("%Y-%m")
    depr_this_month = await db.dewi_asset_depreciation.aggregate([
        {"$match": {"period": current_period}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    depr_amount = depr_this_month[0]["total"] if depr_this_month else 0

    # Warranty & Insurance expiring in 30 days
    from datetime import timedelta
    today_str = date.today().isoformat()
    in_30_days = (date.today() + timedelta(days=30)).isoformat()
    warranty_expiring_count = await db.dewi_assets.count_documents({
        "warranty_expiry_date": {"$gte": today_str, "$lte": in_30_days},
        "status": {"$ne": "disposed"}
    })
    insurance_expiring_count = await db.dewi_assets.count_documents({
        "insurance_expiry_date": {"$gte": today_str, "$lte": in_30_days},
        "status": {"$ne": "disposed"}
    })

    return {
        "summary": {
            "total_assets": total,
            "active": active,
            "in_maintenance": in_maintenance,
            "disposed": disposed,
            "total_purchase_cost": round(totals["total_cost"], 2),
            "total_accumulated_depreciation": round(totals["total_accumulated_depr"], 2),
            "total_nbv": round(nbv, 2),
            "depreciation_this_month": round(depr_amount, 2),
            "warranty_expiring_soon": warranty_expiring_count,
            "insurance_expiring_soon": insurance_expiring_count,
        },
        "by_category": [{"category": b["_id"], "count": b["count"], "total_cost": b["total_cost"]} for b in by_category],
        "recent_assets": [_ser(a) for a in recent],
    }


# ─── Categories ───────────────────────────────────────────────────────────

@router.get("/categories")
async def list_categories(request: Request):
    user = await require_auth(request)
    db = get_db()
    await _ensure_default_categories(db)
    cats = await db.dewi_asset_categories.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return [_ser(c) for c in cats]


@router.post("/categories")
async def create_category(request: Request):
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Nama kategori wajib diisi.")
    doc = {
        "id": _uid(),
        "name": name,
        "code": (body.get("code") or name[:2].upper()).upper(),
        "useful_life_years": int(body.get("useful_life_years") or 5),
        "depr_method": body.get("depr_method", "straight_line"),
        "coa_asset_account": (body.get("coa_asset_account") or "").strip(),
        "coa_depreciation_account": (body.get("coa_depreciation_account") or "").strip(),
        "created_at": _now(),
    }
    await db.dewi_asset_categories.insert_one(doc)
    return _ser(doc)


@router.put("/categories/{cat_id}")
async def update_category(cat_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    update = {}
    if "name" in body:
        update["name"] = body["name"]
    if "code" in body:
        update["code"] = body["code"].upper()
    if "useful_life_years" in body:
        update["useful_life_years"] = int(body["useful_life_years"])
    if "depr_method" in body:
        update["depr_method"] = body["depr_method"]
    if "coa_asset_account" in body:
        update["coa_asset_account"] = (body["coa_asset_account"] or "").strip()
    if "coa_depreciation_account" in body:
        update["coa_depreciation_account"] = (body["coa_depreciation_account"] or "").strip()
    if update:
        await db.dewi_asset_categories.update_one({"id": cat_id}, {"$set": update})
    return {"ok": True}


@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    cnt = await db.dewi_assets.count_documents({"category_id": cat_id})
    if cnt > 0:
        raise HTTPException(400, f"Tidak bisa dihapus: ada {cnt} aset menggunakan kategori ini.")
    await db.dewi_asset_categories.delete_one({"id": cat_id})
    return {"ok": True}


# ─── Assets CRUD ──────────────────────────────────────────────────────────

@router.get("")
async def list_assets(
    request: Request,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    assigned_to: Optional[str] = None,
):
    user = await require_auth(request)
    db = get_db()
    query = {}
    if status:
        query["status"] = status
    if category_id:
        query["category_id"] = category_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"asset_number": {"$regex": search, "$options": "i"}},
            {"serial_number": {"$regex": search, "$options": "i"}},
        ]
    if assigned_to:
        query["assigned_to_id"] = assigned_to

    total = await db.dewi_assets.count_documents(query)
    skip = (page - 1) * limit
    assets = await db.dewi_assets.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {
        "items": [_ser(a) for a in assets],
        "pagination": {"page": page, "page_size": limit, "total": total,
                       "total_pages": math.ceil(total / limit) if total else 1}
    }


@router.post("")
async def create_asset(request: Request):
    user = await require_auth(request)
    db = get_db()
    await _ensure_default_categories(db)
    body = await request.json()
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Nama aset wajib diisi.")
    purchase_cost = float(body.get("purchase_cost") or 0)
    if purchase_cost <= 0:
        raise HTTPException(400, "Harga beli harus lebih dari 0.")

    # Get category
    cat = None
    if body.get("category_id"):
        cat = await db.dewi_asset_categories.find_one({"id": body["category_id"]}, {"_id": 0})
    if not cat:
        cat = {"id": "", "name": "Lain-lain", "code": "LN", "useful_life_years": 5, "depr_method": "straight_line"}

    useful_life_months = int(body.get("useful_life_months") or (cat["useful_life_years"] * 12))
    residual_value = float(body.get("residual_value") or purchase_cost * 0.05)
    depr_method = body.get("depr_method") or cat.get("depr_method", "straight_line")

    asset_number = await _gen_asset_number(db, cat["code"])
    purchase_date = (body.get("purchase_date") or date.today().isoformat())[:10]
    monthly_depr = _calc_straight_line_monthly(purchase_cost, residual_value, useful_life_months)

    doc = {
        "id": _uid(),
        "asset_number": asset_number,
        "name": name,
        "category_id": cat["id"],
        "category_name": cat["name"],
        "purchase_date": purchase_date,
        "purchase_cost": purchase_cost,
        "residual_value": residual_value,
        "useful_life_months": useful_life_months,
        "depreciation_method": depr_method,
        "monthly_depreciation": monthly_depr,
        "accumulated_depreciation": 0.0,
        "location": (body.get("location") or "").strip(),
        "serial_number": (body.get("serial_number") or "").strip(),
        "brand": (body.get("brand") or "").strip(),
        "model": (body.get("model") or "").strip(),
        "department": (body.get("department") or "").strip(),
        "status": "active",
        "assigned_to_id": None,
        "assigned_to_name": None,
        "notes": (body.get("notes") or "").strip(),
        "procurement_request_id": body.get("procurement_request_id"),
        "created_by": user["id"],
        "created_by_name": user.get("name", ""),
        "created_at": _now(),
        "updated_at": _now(),
        "disposed_at": None,
        "journal_purchase_id": None,
        # Warranty fields
        "warranty_expiry_date": (body.get("warranty_expiry_date") or "")[:10] or None,
        "warranty_provider":    (body.get("warranty_provider") or "").strip(),
        "warranty_terms":       (body.get("warranty_terms") or "").strip(),
        # Insurance fields
        "insurance_policy_number": (body.get("insurance_policy_number") or "").strip(),
        "insurance_provider":      (body.get("insurance_provider") or "").strip(),
        "insurance_expiry_date":   (body.get("insurance_expiry_date") or "")[:10] or None,
        "insurance_value":         float(body.get("insurance_value") or 0),
    }

    # Auto-create finance journal for purchase (draft)
    je_id = await _create_finance_journal(
        db, user["id"], user.get("name", ""),
        purchase_date,
        f"Pembelian Aset: {name} ({asset_number})",
        [
            {"account_code": "1500", "account_name": "Aset Tetap", "account_type": "asset",
             "debit": purchase_cost, "credit": 0.0, "description": f"Pembelian {name}"},
            {"account_code": "1100", "account_name": "Kas / Bank", "account_type": "asset",
             "debit": 0.0, "credit": purchase_cost, "description": f"Pembayaran {name}"},
        ],
        source_ref=asset_number
    )
    doc["journal_purchase_id"] = je_id
    await db.dewi_assets.insert_one(doc)
    return _ser(doc)


@router.get("/expiring-alerts")
async def get_expiring_alerts(request: Request, days: int = 30):
    """Aset dengan warranty atau insurance yang akan expired dalam N hari ke depan."""
    user = await require_auth(request)
    db = get_db()
    from datetime import timedelta
    today_str = date.today().isoformat()
    future_str = (date.today() + timedelta(days=days)).isoformat()

    warranty_expiring = await db.dewi_assets.find(
        {"warranty_expiry_date": {"$gte": today_str, "$lte": future_str}, "status": {"$ne": "disposed"}},
        {"_id": 0}
    ).sort("warranty_expiry_date", 1).to_list(50)

    insurance_expiring = await db.dewi_assets.find(
        {"insurance_expiry_date": {"$gte": today_str, "$lte": future_str}, "status": {"$ne": "disposed"}},
        {"_id": 0}
    ).sort("insurance_expiry_date", 1).to_list(50)

    # Also expired (past due)
    warranty_expired = await db.dewi_assets.find(
        {"warranty_expiry_date": {"$ne": None, "$lt": today_str}, "status": {"$ne": "disposed"}},
        {"_id": 0}
    ).sort("warranty_expiry_date", -1).limit(20).to_list(20)

    insurance_expired = await db.dewi_assets.find(
        {"insurance_expiry_date": {"$ne": None, "$lt": today_str}, "status": {"$ne": "disposed"}},
        {"_id": 0}
    ).sort("insurance_expiry_date", -1).limit(20).to_list(20)

    return {
        "warranty_expiring":  [_ser(a) for a in warranty_expiring],
        "warranty_expired":   [_ser(a) for a in warranty_expired],
        "insurance_expiring": [_ser(a) for a in insurance_expiring],
        "insurance_expired":  [_ser(a) for a in insurance_expired],
    }


@router.get("/my-assets")
async def get_my_assets(request: Request):
    """Aset yang ditugaskan ke user saat ini (untuk Portal Saya)."""
    user = await require_auth(request)
    db = get_db()
    assets = await db.dewi_assets.find(
        {"assigned_to_id": user["id"], "status": {"$ne": "disposed"}},
        {"_id": 0}
    ).sort("name", 1).to_list(100)
    return [_ser(a) for a in assets]


@router.post("/bulk-import/preview")
async def bulk_import_preview(request: Request, file: UploadFile = File(...)):
    """Parse CSV/Excel, kembalikan preview baris + kolom untuk column mapping."""
    user = await require_auth(request)
    content = await file.read()
    import io
    try:
        if file.filename.endswith('.csv'):
            import csv
            text = content.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        else:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content), dtype=str).fillna('')
            rows = df.to_dict(orient='records')
    except Exception as e:
        raise HTTPException(400, f"Gagal parse file: {e}")

    if not rows:
        raise HTTPException(400, "File kosong atau tidak ada data.")

    # Return up to 5 preview rows + all column headers
    return {
        "columns": list(rows[0].keys()),
        "preview": rows[:5],
        "total_rows": len(rows),
    }


@router.post("/bulk-import/execute")
async def bulk_import_execute(request: Request):
    """Eksekusi bulk import aset dari data yang sudah di-mapping."""
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    rows = body.get("rows", [])          # list of dicts with asset field names
    category_id = body.get("category_id", "")

    if not rows:
        raise HTTPException(400, "Tidak ada data untuk diimport.")
    if not category_id:
        raise HTTPException(400, "Pilih kategori aset terlebih dahulu.")

    # Get category
    cat = await db.dewi_asset_categories.find_one({"id": category_id})
    if not cat:
        raise HTTPException(404, "Kategori tidak ditemukan.")

    created = []
    errors = []
    for i, row in enumerate(rows):
        try:
            # Validate required: name, purchase_date, purchase_cost
            name = str(row.get("name", "")).strip()
            if not name:
                errors.append({"row": i + 1, "error": "Nama aset wajib diisi"})
                continue
            purchase_date_str = str(row.get("purchase_date", "")).strip()
            if not purchase_date_str:
                purchase_date_str = date.today().isoformat()
            try:
                pd_parsed = date.fromisoformat(purchase_date_str[:10])
                purchase_date = pd_parsed.isoformat()
            except Exception:
                purchase_date = date.today().isoformat()

            purchase_cost = 0.0
            try:
                raw_cost = str(row.get("purchase_cost", "0")).replace(",", "").strip()
                purchase_cost = float(raw_cost) if raw_cost else 0.0
            except Exception:
                pass

            useful_life_months = 0
            try:
                raw_ul = str(row.get("useful_life_months", "")).strip()
                useful_life_months = int(float(raw_ul)) if raw_ul else (cat.get("useful_life_years", 5) * 12)
            except Exception:
                useful_life_months = (cat.get("useful_life_years", 5) * 12)

            residual_value = 0.0
            try:
                raw_rv = str(row.get("residual_value", "0")).replace(",", "").strip()
                residual_value = float(raw_rv) if raw_rv else 0.0
            except Exception:
                pass

            asset_number = await _gen_asset_number(db, cat["code"])
            doc = {
                "id":              _uid(),
                "asset_number":    asset_number,
                "name":            name,
                "category_id":     cat["id"],
                "category_name":   cat["name"],
                "purchase_date":   purchase_date,
                "purchase_cost":   purchase_cost,
                "residual_value":  residual_value,
                "useful_life_months": useful_life_months,
                "depreciation_method": cat.get("depr_method", "straight_line"),
                "accumulated_depreciation": 0.0,
                "status":          "active",
                "location":        str(row.get("location", "")).strip(),
                "department":      str(row.get("department", "")).strip(),
                "serial_number":   str(row.get("serial_number", "")).strip(),
                "brand":           str(row.get("brand", "")).strip(),
                "model":           str(row.get("model", "")).strip(),
                "notes":           str(row.get("notes", "")).strip(),
                "warranty_expiry_date":  str(row.get("warranty_expiry_date", ""))[:10] or None,
                "warranty_provider":     str(row.get("warranty_provider", "")).strip(),
                "warranty_terms":        str(row.get("warranty_terms", "")).strip(),
                "insurance_policy_number": str(row.get("insurance_policy_number", "")).strip(),
                "insurance_provider":    str(row.get("insurance_provider", "")).strip(),
                "insurance_expiry_date": str(row.get("insurance_expiry_date", ""))[:10] or None,
                "insurance_value":       float(row.get("insurance_value", 0) or 0),
                "photo_url":       None,
                "assigned_to":     None,
                "assigned_to_name": "",
                "assigned_at":     None,
                "procurement_request_id": None,
                "created_by":      user["id"],
                "created_by_name": user.get("name", ""),
                "created_at":      _now(),
                "updated_at":      _now(),
                "disposed_at":     None,
                "journal_purchase_id": None,
            }
            await db.dewi_assets.insert_one(doc)
            created.append(asset_number)
        except Exception as e:
            errors.append({"row": i + 1, "error": str(e)})

    return {
        "ok": True,
        "created_count": len(created),
        "error_count": len(errors),
        "created": created,
        "errors": errors[:10],   # cap at 10 error samples
    }


@router.get("/bulk-import/template")
async def bulk_import_template(request: Request):
    """Download template Excel untuk bulk import aset."""
    from fastapi.responses import StreamingResponse
    import io
    import openpyxl
    user = await require_auth(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Aset"
    headers = [
        "name*", "purchase_date*", "purchase_cost*",
        "useful_life_months", "residual_value",
        "serial_number", "brand", "model", "location", "department", "notes",
        "warranty_expiry_date", "warranty_provider", "warranty_terms",
        "insurance_policy_number", "insurance_provider", "insurance_expiry_date", "insurance_value",
    ]
    ws.append(headers)
    # Sample row
    ws.append([
        "Laptop Dell XPS 15", "2026-01-01", "15000000",
        "48", "0",
        "SN-DELL-0001", "Dell", "XPS 15", "Kantor Pusat", "IT", "Laptop kerja",
        "2028-01-01", "Dell Support", "On-site 2 tahun",
        "POL-2026-001", "Jasindo", "2027-01-01", "20000000",
    ])
    # Style header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_aset.xlsx"},
    )


@router.post("/bulk-import/execute-file")
async def bulk_import_execute_file(
    request: Request,
    file: UploadFile = File(...),
):
    """Eksekusi bulk import dengan upload file + column mapping JSON."""
    from fastapi import Form as FastForm
    import json as _json
    import io as _io
    user = await require_auth(request)
    db = get_db()

    # Read multipart form fields manually
    form = await request.form()
    mapping_json = form.get("mapping", "{}")
    category_id = form.get("category_id", "")
    file_obj = form.get("file")

    try:
        mapping = _json.loads(mapping_json)
    except Exception:
        raise HTTPException(400, "mapping JSON tidak valid")

    if not category_id:
        raise HTTPException(400, "category_id wajib")

    cat = await db.dewi_asset_categories.find_one({"id": category_id})
    if not cat:
        raise HTTPException(404, "Kategori tidak ditemukan.")

    # Read file content
    content = await file_obj.read()
    filename = file_obj.filename or ""

    try:
        if filename.endswith('.csv'):
            import csv
            text = content.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(_io.StringIO(text))
            all_rows = [dict(r) for r in reader]
        else:
            import pandas as pd
            df = pd.read_excel(_io.BytesIO(content), dtype=str).fillna('')
            all_rows = df.to_dict(orient='records')
    except Exception as e:
        raise HTTPException(400, f"Gagal parse file: {e}")

    if not all_rows:
        raise HTTPException(400, "File kosong atau tidak ada data.")

    # Apply mapping: transform raw column names → asset field names
    def apply_mapping(row):
        out = {}
        for field, col in mapping.items():
            if col and col in row:
                out[field] = row[col]
        return out

    created = []
    errors = []
    for i, raw in enumerate(all_rows):
        mapped = apply_mapping(raw)
        try:
            name = str(mapped.get("name", "")).strip()
            if not name:
                errors.append({"row": i + 1, "error": "Nama aset kosong"})
                continue

            purchase_date_str = str(mapped.get("purchase_date", "")).strip()
            try:
                purchase_date = date.fromisoformat(purchase_date_str[:10]).isoformat()
            except Exception:
                purchase_date = date.today().isoformat()

            try:
                purchase_cost = float(str(mapped.get("purchase_cost", "0")).replace(",", "") or "0")
            except Exception:
                purchase_cost = 0.0

            try:
                ul = str(mapped.get("useful_life_months", "")).strip()
                useful_life_months = int(float(ul)) if ul else cat.get("useful_life_years", 5) * 12
            except Exception:
                useful_life_months = cat.get("useful_life_years", 5) * 12

            try:
                rv = str(mapped.get("residual_value", "0")).replace(",", "") or "0"
                residual_value = float(rv)
            except Exception:
                residual_value = 0.0

            asset_number = await _gen_asset_number(db, cat["code"])
            doc = {
                "id":              _uid(),
                "asset_number":    asset_number,
                "name":            name,
                "category_id":     cat["id"],
                "category_name":   cat["name"],
                "purchase_date":   purchase_date,
                "purchase_cost":   purchase_cost,
                "residual_value":  residual_value,
                "useful_life_months": useful_life_months,
                "depreciation_method": cat.get("depr_method", "straight_line"),
                "accumulated_depreciation": 0.0,
                "status":          "active",
                "location":        str(mapped.get("location", "")).strip(),
                "department":      str(mapped.get("department", "")).strip(),
                "serial_number":   str(mapped.get("serial_number", "")).strip(),
                "brand":           str(mapped.get("brand", "")).strip(),
                "model":           str(mapped.get("model", "")).strip(),
                "notes":           str(mapped.get("notes", "")).strip(),
                "warranty_expiry_date":  (str(mapped.get("warranty_expiry_date", ""))[:10] or None),
                "warranty_provider":     str(mapped.get("warranty_provider", "")).strip(),
                "warranty_terms":        str(mapped.get("warranty_terms", "")).strip(),
                "insurance_policy_number": str(mapped.get("insurance_policy_number", "")).strip(),
                "insurance_provider":    str(mapped.get("insurance_provider", "")).strip(),
                "insurance_expiry_date": (str(mapped.get("insurance_expiry_date", ""))[:10] or None),
                "insurance_value":       float(mapped.get("insurance_value", 0) or 0),
                "photo_url":       None,
                "assigned_to":     None,
                "assigned_to_name": "",
                "assigned_at":     None,
                "procurement_request_id": None,
                "created_by":      user["id"],
                "created_by_name": user.get("name", ""),
                "created_at":      _now(),
                "updated_at":      _now(),
                "disposed_at":     None,
                "journal_purchase_id": None,
            }
            await db.dewi_assets.insert_one(doc)
            created.append(asset_number)
        except Exception as e:
            errors.append({"row": i + 1, "error": str(e)})

    return {
        "ok": True,
        "created_count": len(created),
        "error_count": len(errors),
        "created": created[:50],
        "errors": errors[:10],
    }


@router.post("/batch-depreciate/{period}")
async def batch_depreciate(period: str, request: Request):
    """Posting depresiasi massal untuk semua aset aktif pada periode YYYY-MM. Idempotent per aset."""
    user = await require_auth(request)
    db = get_db()
    assets = await db.dewi_assets.find(
        {"status": {"$ne": "disposed"}}, {"_id": 0}
    ).to_list(1000)
    results = {"posted": [], "skipped": [], "errors": []}
    for asset in assets:
        asset_id = asset["id"]
        existing = await db.dewi_asset_depreciation.find_one({"asset_id": asset_id, "period": period})
        if existing:
            results["skipped"].append({"id": asset_id, "number": asset.get("asset_number"), "reason": "Sudah diposting"})
            continue
        cost = float(asset.get("purchase_cost", 0))
        residual = float(asset.get("residual_value", 0))
        accum = float(asset.get("accumulated_depreciation", 0))
        nbv = cost - accum
        if nbv <= residual:
            results["skipped"].append({"id": asset_id, "number": asset.get("asset_number"), "reason": "Sudah habis"})
            continue
        try:
            monthly = float(asset.get("monthly_depreciation") or
                            _calc_straight_line_monthly(cost, residual, asset.get("useful_life_months", 60)))
            depr_amount = min(monthly, nbv - residual)
            new_accum = accum + depr_amount
            je_id = await _create_finance_journal(
                db, user["id"], user.get("name", ""),
                f"{period}-28",
                f"Depresiasi Massal: {asset['name']} ({asset.get('asset_number', '')}) - {period}",
                [
                    {"account_code": "6200", "account_name": "Beban Depresiasi", "account_type": "expense",
                     "debit": depr_amount, "credit": 0.0, "description": f"Depresiasi {asset.get('asset_number', '')}"},
                    {"account_code": "1590", "account_name": "Akumulasi Depresiasi", "account_type": "asset",
                     "debit": 0.0, "credit": depr_amount, "description": f"Akumulasi {asset.get('asset_number', '')}"},
                ],
                source_ref=asset.get("asset_number", "")
            )
            depr_doc = {
                "id": _uid(), "asset_id": asset_id, "asset_number": asset.get("asset_number", ""),
                "asset_name": asset["name"], "period": period,
                "amount": depr_amount, "cumulative": new_accum,
                "nbv_before": nbv, "nbv_after": cost - new_accum,
                "journal_id": je_id, "created_by": user["id"], "created_at": _now(),
            }
            await db.dewi_asset_depreciation.insert_one(depr_doc)
            await db.dewi_assets.update_one(
                {"id": asset_id}, {"$set": {"accumulated_depreciation": new_accum, "updated_at": _now()}}
            )
            results["posted"].append({"id": asset_id, "number": asset.get("asset_number"), "amount": depr_amount})
        except Exception as e:
            results["errors"].append({"id": asset_id, "number": asset.get("asset_number"), "error": str(e)})
    return {
        "period": period,
        "total_posted": len(results["posted"]),
        "total_skipped": len(results["skipped"]),
        "total_errors": len(results["errors"]),
        "details": results
    }




@router.get("/disposal-requests")
async def list_disposal_requests_alias(request: Request, status: str = "pending"):
    """List permintaan disposal (alias sebelum /{asset_id} agar routing benar)."""
    return await list_disposal_requests(request, status)


@router.get("/{asset_id}")
async def get_asset(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")

    # Calculate schedule
    cost = float(asset.get("purchase_cost", 0))
    residual = float(asset.get("residual_value", 0))
    months = int(asset.get("useful_life_months", 60))
    accum = float(asset.get("accumulated_depreciation", 0))
    nbv = cost - accum

    # Recent depreciation
    depr_history = await db.dewi_asset_depreciation.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("period", -1).limit(12).to_list(12)

    asset_out = _ser(asset)
    asset_out["nbv"] = round(nbv, 2)
    asset_out["depreciation_history"] = [_ser(d) for d in depr_history]
    asset_out["fully_depreciated"] = nbv <= residual
    return asset_out


@router.put("/{asset_id}")
async def update_asset(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    body = await request.json()
    allowed = [
        "name", "location", "serial_number", "brand", "model",
        "department", "notes", "status",
        # Warranty
        "warranty_expiry_date", "warranty_provider", "warranty_terms",
        # Insurance
        "insurance_policy_number", "insurance_provider", "insurance_expiry_date", "insurance_value",
    ]
    update = {k: body[k] for k in allowed if k in body}
    if update:
        update["updated_at"] = _now()
        await db.dewi_assets.update_one({"id": asset_id}, {"$set": update})
    return {"ok": True}


# ─── Assignment ────────────────────────────────────────────────────────────

@router.post("/{asset_id}/assign")
async def assign_asset(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    body = await request.json()
    assign_to_id = body.get("user_id") or ""
    assign_to_name = body.get("user_name") or ""
    if not assign_to_id:
        raise HTTPException(400, "ID karyawan wajib diisi.")

    # Create assignment record
    assn_doc = {
        "id": _uid(),
        "asset_id": asset_id,
        "asset_name": asset["name"],
        "asset_number": asset["asset_number"],
        "assigned_to_id": assign_to_id,
        "assigned_to_name": assign_to_name,
        "assigned_by_id": user["id"],
        "assigned_by_name": user.get("name", ""),
        "assigned_date": (body.get("assigned_date") or date.today().isoformat())[:10],
        "returned_date": None,
        "notes": (body.get("notes") or "").strip(),
        "status": "active",
        "created_at": _now(),
    }
    await db.dewi_asset_assignments.insert_one(assn_doc)
    await db.dewi_assets.update_one(
        {"id": asset_id},
        {"$set": {"assigned_to_id": assign_to_id, "assigned_to_name": assign_to_name, "updated_at": _now()}}
    )
    return _ser(assn_doc)


@router.post("/{asset_id}/unassign")
async def unassign_asset(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    today = date.today().isoformat()
    # Update active assignment
    await db.dewi_asset_assignments.update_many(
        {"asset_id": asset_id, "status": "active"},
        {"$set": {"status": "returned", "returned_date": today}}
    )
    await db.dewi_assets.update_one(
        {"id": asset_id},
        {"$set": {"assigned_to_id": None, "assigned_to_name": None, "updated_at": _now()}}
    )
    return {"ok": True}


@router.get("/{asset_id}/assignments")
async def get_asset_assignments(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    recs = await db.dewi_asset_assignments.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("assigned_date", -1).to_list(100)
    return [_ser(r) for r in recs]


# ─── Maintenance ──────────────────────────────────────────────────────────

@router.post("/{asset_id}/maintenance")
async def add_maintenance(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    body = await request.json()
    doc = {
        "id": _uid(),
        "asset_id": asset_id,
        "asset_name": asset["name"],
        "type": body.get("type", "corrective"),  # scheduled | corrective | preventive
        "description": (body.get("description") or "").strip(),
        "cost": float(body.get("cost") or 0),
        "performed_by": (body.get("performed_by") or "").strip(),
        "maintenance_date": (body.get("maintenance_date") or date.today().isoformat())[:10],
        "next_scheduled": body.get("next_scheduled"),
        "status": body.get("status", "completed"),
        "notes": (body.get("notes") or "").strip(),
        "created_by": user["id"],
        "created_at": _now(),
    }
    await db.dewi_asset_maintenance.insert_one(doc)
    # Update asset status if in_maintenance
    if doc["status"] == "in_progress":
        await db.dewi_assets.update_one({"id": asset_id}, {"$set": {"status": "in_maintenance", "updated_at": _now()}})
    elif doc["status"] == "completed" and asset.get("status") == "in_maintenance":
        await db.dewi_assets.update_one({"id": asset_id}, {"$set": {"status": "active", "updated_at": _now()}})
    return _ser(doc)


@router.get("/{asset_id}/maintenance")
async def get_maintenance_history(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    recs = await db.dewi_asset_maintenance.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("maintenance_date", -1).to_list(100)
    return [_ser(r) for r in recs]


# ─── Depreciation ─────────────────────────────────────────────────────────

@router.post("/{asset_id}/depreciate/{period}")
async def post_depreciation(asset_id: str, period: str, request: Request):
    """Post depresiasi untuk 1 aset pada periode YYYY-MM. Idempotent."""
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    if asset.get("status") == "disposed":
        raise HTTPException(400, "Aset sudah dilepas, tidak bisa depresiasi.")

    # Idempotency check
    existing = await db.dewi_asset_depreciation.find_one({"asset_id": asset_id, "period": period})
    if existing:
        raise HTTPException(400, f"Depresiasi periode {period} sudah diposting.")

    cost = float(asset["purchase_cost"])
    residual = float(asset.get("residual_value", 0))
    accum = float(asset.get("accumulated_depreciation", 0))
    nbv = cost - accum

    if nbv <= residual:
        raise HTTPException(400, "Aset sudah habis didepresiasi (NBV = nilai residu).")

    monthly = float(asset.get("monthly_depreciation") or
                    _calc_straight_line_monthly(cost, residual, asset.get("useful_life_months", 60)))
    depr_amount = min(monthly, nbv - residual)
    new_accum = accum + depr_amount

    period_date = f"{period}-28"  # End of period approximate
    je_id = await _create_finance_journal(
        db, user["id"], user.get("name", ""),
        period_date,
        f"Depresiasi Aset: {asset['name']} ({asset['asset_number']}) - {period}",
        [
            {"account_code": "6200", "account_name": "Beban Depresiasi", "account_type": "expense",
             "debit": depr_amount, "credit": 0.0, "description": f"Depresiasi {asset['asset_number']}"},
            {"account_code": "1590", "account_name": "Akumulasi Depresiasi", "account_type": "asset",
             "debit": 0.0, "credit": depr_amount, "description": f"Akumulasi {asset['asset_number']}"},
        ],
        source_ref=asset["asset_number"]
    )

    depr_doc = {
        "id": _uid(),
        "asset_id": asset_id,
        "asset_number": asset["asset_number"],
        "asset_name": asset["name"],
        "period": period,
        "amount": depr_amount,
        "cumulative": new_accum,
        "nbv_before": nbv,
        "nbv_after": cost - new_accum,
        "journal_id": je_id,
        "created_by": user["id"],
        "created_at": _now(),
    }
    await db.dewi_asset_depreciation.insert_one(depr_doc)
    await db.dewi_assets.update_one(
        {"id": asset_id},
        {"$set": {"accumulated_depreciation": new_accum, "updated_at": _now()}}
    )
    return _ser(depr_doc)


@router.get("/{asset_id}/depreciation-history")
async def get_depreciation_history(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    recs = await db.dewi_asset_depreciation.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("period", -1).to_list(100)
    return [_ser(r) for r in recs]


# ─── Disposal ─────────────────────────────────────────────────────────────

@router.post("/{asset_id}/dispose")
async def dispose_asset(asset_id: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    if asset.get("status") == "disposed":
        raise HTTPException(400, "Aset sudah dilepas.")
    body = await request.json()
    disposal_date = (body.get("disposal_date") or date.today().isoformat())[:10]
    disposal_value = float(body.get("disposal_value") or 0)
    reason = (body.get("reason") or "").strip()

    cost = float(asset["purchase_cost"])
    accum = float(asset.get("accumulated_depreciation", 0))
    nbv = cost - accum
    gain_loss = disposal_value - nbv

    # Finance journal for disposal
    lines = [
        {"account_code": "1590", "account_name": "Akumulasi Depresiasi", "account_type": "asset",
         "debit": accum, "credit": 0.0, "description": f"Hapus akumulasi {asset['asset_number']}"},
        {"account_code": "1500", "account_name": "Aset Tetap", "account_type": "asset",
         "debit": 0.0, "credit": cost, "description": f"Hapus aset {asset['asset_number']}"},
    ]
    if disposal_value > 0:
        lines.append({"account_code": "1100", "account_name": "Kas", "account_type": "asset",
                      "debit": disposal_value, "credit": 0.0, "description": "Penerimaan disposal"})
    if gain_loss > 0:
        lines.append({"account_code": "8100", "account_name": "Keuntungan Disposal Aset", "account_type": "revenue",
                      "debit": 0.0, "credit": gain_loss, "description": "Keuntungan disposal"})
    elif gain_loss < 0:
        lines.append({"account_code": "6300", "account_name": "Kerugian Disposal Aset", "account_type": "expense",
                      "debit": abs(gain_loss), "credit": 0.0, "description": "Kerugian disposal"})

    je_id = await _create_finance_journal(
        db, user["id"], user.get("name", ""),
        disposal_date,
        f"Disposal Aset: {asset['name']} ({asset['asset_number']})",
        lines,
        source_ref=asset["asset_number"]
    )
    await db.dewi_assets.update_one(
        {"id": asset_id},
        {"$set": {
            "status": "disposed",
            "disposed_at": _now(),
            "disposal_value": disposal_value,
            "disposal_reason": reason,
            "disposal_journal_id": je_id,
            "updated_at": _now()
        }}
    )
    return {"ok": True, "gain_loss": gain_loss, "journal_id": je_id}


# ─── Asset Disposal Approval Workflow ─────────────────────────────────────────
# Aset dengan NBV > DISPOSAL_APPROVAL_THRESHOLD memerlukan approval sebelum dispose

DISPOSAL_APPROVAL_THRESHOLD = 5_000_000  # 5 juta IDR


@router.post("/{asset_id}/request-disposal")
async def request_disposal(asset_id: str, request: Request):
    """Buat permintaan disposal untuk aset bernilai tinggi (NBV > threshold)."""
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    if asset.get("status") == "disposed":
        raise HTTPException(400, "Aset sudah dilepas.")

    cost = float(asset.get("purchase_cost", 0))
    accum = float(asset.get("accumulated_depreciation", 0))
    nbv = cost - accum

    # Check if there's already a pending request
    existing = await db.dewi_asset_disposal_requests.find_one(
        {"asset_id": asset_id, "status": "pending"}
    )
    if existing:
        raise HTTPException(400, "Sudah ada permintaan disposal yang menunggu approval.")

    body = await request.json()
    disposal_date  = (body.get("disposal_date") or date.today().isoformat())[:10]
    disposal_value = float(body.get("disposal_value") or 0)
    reason         = (body.get("reason") or "").strip()

    if not reason:
        raise HTTPException(400, "Alasan disposal wajib diisi.")

    doc = {
        "id":              _uid(),
        "asset_id":        asset_id,
        "asset_number":    asset.get("asset_number", ""),
        "asset_name":      asset.get("name", ""),
        "nbv":             round(nbv, 2),
        "disposal_date":   disposal_date,
        "disposal_value":  disposal_value,
        "reason":          reason,
        "status":          "pending",   # pending | approved | rejected
        "requested_by":    user["id"],
        "requested_by_name": user.get("name", ""),
        "requested_at":    _now(),
        "reviewed_by":     None,
        "reviewed_by_name": "",
        "reviewed_at":     None,
        "review_notes":    "",
        "journal_id":      None,
    }
    await db.dewi_asset_disposal_requests.insert_one(doc)
    # Mark asset as pending_disposal so it shows up in status
    await db.dewi_assets.update_one(
        {"id": asset_id},
        {"$set": {"status": "pending_disposal", "updated_at": _now()}}
    )
    return _ser(doc)


@router.get("/disposal-requests")
async def list_disposal_requests(request: Request, status: str = "pending"):
    """List permintaan disposal. Status: pending | approved | rejected | all."""
    user = await require_auth(request)
    db = get_db()
    filt = {}
    if status != "all":
        filt["status"] = status
    reqs = await db.dewi_asset_disposal_requests.find(filt, {"_id": 0}).sort("requested_at", -1).to_list(100)
    return [_ser(r) for r in reqs]


@router.patch("/disposal-requests/{req_id}/approve")
async def approve_disposal_request(req_id: str, request: Request):
    """Approve permintaan disposal → eksekusi dispose + jurnal Finance."""
    user = await require_auth(request)
    if user.get("role") not in ("admin", "superadmin", "finance", "manager"):
        raise HTTPException(403, "Hanya admin/finance/manager yang bisa approve disposal.")
    db = get_db()
    req = await db.dewi_asset_disposal_requests.find_one({"id": req_id})
    if not req:
        raise HTTPException(404, "Request tidak ditemukan.")
    if req["status"] != "pending":
        raise HTTPException(400, f"Request sudah {req['status']}.")

    body = await request.json()
    notes = (body.get("notes") or "").strip()

    asset = await db.dewi_assets.find_one({"id": req["asset_id"]})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")

    # Execute disposal (same logic as dispose_asset)
    cost   = float(asset["purchase_cost"])
    accum  = float(asset.get("accumulated_depreciation", 0))
    nbv    = cost - accum
    disposal_value = float(req.get("disposal_value", 0))
    gain_loss      = disposal_value - nbv

    lines = [
        {"account_code": "1590", "account_name": "Akumulasi Depresiasi", "account_type": "asset",
         "debit": accum, "credit": 0.0, "description": f"Hapus akumulasi {asset['asset_number']}"},
        {"account_code": "1500", "account_name": "Aset Tetap", "account_type": "asset",
         "debit": 0.0, "credit": cost, "description": f"Hapus aset {asset['asset_number']}"},
    ]
    if disposal_value > 0:
        lines.append({"account_code": "1100", "account_name": "Kas", "account_type": "asset",
                      "debit": disposal_value, "credit": 0.0, "description": "Penerimaan disposal"})
    if gain_loss > 0:
        lines.append({"account_code": "8100", "account_name": "Keuntungan Disposal Aset", "account_type": "revenue",
                      "debit": 0.0, "credit": gain_loss, "description": "Keuntungan disposal"})
    elif gain_loss < 0:
        lines.append({"account_code": "6300", "account_name": "Kerugian Disposal Aset", "account_type": "expense",
                      "debit": abs(gain_loss), "credit": 0.0, "description": "Kerugian disposal"})

    je_id = await _create_finance_journal(
        db, user["id"], user.get("name", ""),
        req["disposal_date"],
        f"Disposal Aset (Approved): {asset['name']} ({asset['asset_number']})",
        lines, source_ref=asset["asset_number"]
    )

    now = _now()
    await db.dewi_assets.update_one(
        {"id": asset["id"]},
        {"$set": {
            "status": "disposed", "disposed_at": now,
            "disposal_value": disposal_value,
            "disposal_reason": req.get("reason", ""),
            "disposal_journal_id": je_id,
            "updated_at": now,
        }}
    )
    await db.dewi_asset_disposal_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": "approved", "reviewed_by": user["id"],
            "reviewed_by_name": user.get("name", ""),
            "reviewed_at": now, "review_notes": notes, "journal_id": je_id,
        }}
    )
    return {"ok": True, "gain_loss": round(gain_loss, 2), "journal_id": je_id}


@router.patch("/disposal-requests/{req_id}/reject")
async def reject_disposal_request(req_id: str, request: Request):
    """Reject permintaan disposal → kembalikan status aset ke active."""
    user = await require_auth(request)
    if user.get("role") not in ("admin", "superadmin", "finance", "manager"):
        raise HTTPException(403, "Hanya admin/finance/manager yang bisa reject disposal.")
    db = get_db()
    req = await db.dewi_asset_disposal_requests.find_one({"id": req_id})
    if not req:
        raise HTTPException(404, "Request tidak ditemukan.")
    if req["status"] != "pending":
        raise HTTPException(400, f"Request sudah {req['status']}.")

    body = await request.json()
    notes = (body.get("notes") or "").strip()

    now = _now()
    await db.dewi_asset_disposal_requests.update_one(
        {"id": req_id},
        {"$set": {
            "status": "rejected", "reviewed_by": user["id"],
            "reviewed_by_name": user.get("name", ""),
            "reviewed_at": now, "review_notes": notes,
        }}
    )
    # Restore asset status to active
    await db.dewi_assets.update_one(
        {"id": req["asset_id"]},
        {"$set": {"status": "active", "updated_at": now}}
    )
    return {"ok": True}


# ─── Init indexes ─────────────────────────────────────────────────────────

# ─── Asset Photo Upload ───────────────────────────────────────────────────
from fastapi import UploadFile, File

@router.post("/{asset_id}/upload-photo")
async def upload_asset_photo(asset_id: str, request: Request, file: UploadFile = File(...)):
    """Upload foto asset untuk visual identification."""
    user = await require_auth(request)
    db = get_db()
    
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0, "asset_number": 1})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    
    # Validate image
    if file.content_type and not file.content_type.startswith('image/'):
        raise HTTPException(400, "File harus berupa gambar (jpg, png, etc)")
    
    if file.size and file.size > 5 * 1024 * 1024:  # 5MB max
        raise HTTPException(400, "Ukuran foto maksimal 5 MB")
    
    # Read and store
    content_bytes = await file.read()
    storage_path = generate_storage_path(f"assets/{asset_id}", file.filename)
    stored = put_object(storage_path, content_bytes, file.content_type or "image/jpeg")
    photo_url = stored["url"]  # Extract URL string from storage result dict
    
    # Update asset
    await db.dewi_assets.update_one(
        {"id": asset_id},
        {"$set": {"photo_url": photo_url, "updated_at": _now()}}
    )
    
    return {"ok": True, "photo_url": photo_url}


# ─── Asset Transfer Workflow ─────────────────────────────────────────────
@router.post("/{asset_id}/transfer")
async def transfer_asset(asset_id: str, request: Request):
    """Transfer asset to new location/department/employee dengan audit trail."""
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    
    # Prepare transfer record
    transfer_doc = {
        "id": _uid(),
        "asset_id": asset_id,
        "asset_number": asset["asset_number"],
        "asset_name": asset["name"],
        "from_location": asset.get("location", ""),
        "from_department": asset.get("department", ""),
        "from_employee_id": asset.get("assigned_to_id", ""),
        "from_employee_name": asset.get("assigned_to_name", ""),
        "to_location": (body.get("to_location") or "").strip(),
        "to_department": (body.get("to_department") or "").strip(),
        "to_employee_id": (body.get("to_employee_id") or "").strip(),
        "to_employee_name": (body.get("to_employee_name") or "").strip(),
        "transfer_date": body.get("transfer_date", _now().isoformat()),
        "reason": (body.get("reason") or "").strip(),
        "notes": (body.get("notes") or "").strip(),
        "transferred_by": user["id"],
        "transferred_by_name": user.get("name", ""),
        "created_at": _now(),
        "status": "completed",
    }
    
    # Insert transfer record
    await db.dewi_asset_transfers.insert_one(transfer_doc)
    
    # Update asset
    update = {"updated_at": _now()}
    if body.get("to_location"):
        update["location"] = body["to_location"]
    if body.get("to_department"):
        update["department"] = body["to_department"]
    if body.get("to_employee_id"):
        update["assigned_to_id"] = body["to_employee_id"]
        update["assigned_to_name"] = body.get("to_employee_name", "")
    
    await db.dewi_assets.update_one({"id": asset_id}, {"$set": update})
    
    return {"ok": True, "transfer_id": transfer_doc["id"]}


@router.get("/{asset_id}/transfer-history")
async def get_transfer_history(asset_id: str, request: Request):
    """Riwayat transfer asset untuk audit trail."""
    await require_auth(request)
    db = get_db()
    transfers = await db.dewi_asset_transfers.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)
    return [_ser(t) for t in transfers]


# ─── Asset Scanning & Location Tracking ───────────────────────────────────
@router.post("/{asset_id}/scan")
async def scan_asset(asset_id: str, request: Request):
    """Scan asset untuk tracking lokasi/movement. Used by WMS/Inventory apps."""
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    
    # Record scan event
    scan_doc = {
        "id": _uid(),
        "asset_id": asset_id,
        "asset_number": asset["asset_number"],
        "asset_name": asset["name"],
        "scanned_by": user["id"],
        "scanned_by_name": user.get("name", ""),
        "scan_type": body.get("scan_type", "location_check"),  # location_check | movement | opname
        "location": (body.get("location") or "").strip(),
        "notes": (body.get("notes") or "").strip(),
        "scanned_at": _now(),
    }
    await db.dewi_asset_scans.insert_one(scan_doc)
    
    # Update asset location if provided
    update = {}
    if body.get("location"):
        update["location"] = body["location"]
        update["updated_at"] = _now()
    if update:
        await db.dewi_assets.update_one({"id": asset_id}, {"$set": update})
    
    return {"ok": True, "scan_id": scan_doc["id"], "asset": _ser(asset)}


@router.get("/scan-by-number/{asset_number}")
async def get_asset_by_number(asset_number: str, request: Request):
    """Resolve asset by asset_number (untuk scanner apps)."""
    await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one(
        {"asset_number": {"$regex": f"^{asset_number}$", "$options": "i"}}, 
        {"_id": 0}
    )
    if not asset:
        raise HTTPException(404, f"Aset dengan nomor '{asset_number}' tidak ditemukan.")
    return _ser(asset)


@router.get("/{asset_id}/scan-history")
async def get_asset_scan_history(asset_id: str, request: Request):
    """Riwayat scan asset (untuk audit trail)."""
    await require_auth(request)
    db = get_db()
    scans = await db.dewi_asset_scans.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("scanned_at", -1).limit(50).to_list(50)
    return [_ser(s) for s in scans]


# ─── Barcode & QR Code Generation ─────────────────────────────────────────
from fastapi.responses import Response

@router.get("/{asset_id}/barcode")
async def get_asset_barcode(asset_id: str, request: Request):
    """Generate Code128 barcode image (PNG) untuk asset_number."""
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0, "asset_number": 1, "name": 1})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    
    try:
        import barcode
        from barcode.writer import ImageWriter
        import io
        
        code_val = asset["asset_number"]
        bc_cls = barcode.get_barcode_class("code128")
        bc_obj = bc_cls(code_val, writer=ImageWriter())
        buf = io.BytesIO()
        bc_obj.write(buf, options={"write_text": True, "font_size": 12, "text_distance": 3, "module_height": 12})
        buf.seek(0)
        return Response(content=buf.read(), media_type="image/png")
    except Exception as e:
        logger.error(f"Barcode generation error: {e}")
        raise HTTPException(500, "Gagal generate barcode")


@router.get("/{asset_id}/qrcode")
async def get_asset_qrcode(asset_id: str, request: Request):
    """Generate QR code image (PNG) dengan JSON lengkap + URL link."""
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    
    try:
        import qrcode
        import json
        import io
        from PIL import Image
        
        # QR payload: JSON lengkap + URL
        base_url = request.base_url.scheme + "://" + request.base_url.netloc
        qr_data = {
            "type": "asset",
            "asset_id": asset["id"],
            "asset_number": asset["asset_number"],
            "name": asset["name"],
            "category": asset.get("category_name", ""),
            "location": asset.get("location", ""),
            "url": f"{base_url}/asset/{asset['id']}"
        }
        
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return Response(content=buf.read(), media_type="image/png")
    except Exception as e:
        logger.error(f"QR code generation error: {e}")
        raise HTTPException(500, "Gagal generate QR code")


@router.get("/{asset_id}/label-pdf")
async def get_asset_label_pdf(asset_id: str, request: Request, template: str = "standard"):
    """Generate printable label PDF (barcode + QR + asset info)."""
    user = await require_auth(request)
    db = get_db()
    asset = await db.dewi_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(404, "Aset tidak ditemukan.")
    
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        import barcode
        from barcode.writer import ImageWriter
        import qrcode
        import json
        import io
        
        buf = io.BytesIO()
        
        # Template sizes
        templates = {
            "standard": (90 * mm, 50 * mm),  # 90x50mm landscape
            "sticker": (50 * mm, 25 * mm),   # 50x25mm small sticker
            "a4": (210 * mm, 297 * mm),       # A4 full page
        }
        page_size = templates.get(template, templates["standard"])
        
        c = rl_canvas.Canvas(buf, pagesize=page_size)
        LW, LH = page_size
        
        if template == "sticker":
            # Small sticker: QR code only
            base_url = request.base_url.scheme + "://" + request.base_url.netloc
            qr_data = json.dumps({
                "type": "asset",
                "asset_id": asset["id"],
                "asset_number": asset["asset_number"],
                "url": f"{base_url}/asset/{asset['id']}"
            })
            qr_obj = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=5, border=1)
            qr_obj.add_data(qr_data)
            qr_obj.make(fit=True)
            qr_img = qr_obj.make_image(fill_color="black", back_color="white")
            qr_buf = io.BytesIO()
            qr_img.save(qr_buf, format="PNG")
            qr_buf.seek(0)
            c.drawImage(ImageReader(qr_buf), 2 * mm, 10 * mm, width=20 * mm, height=20 * mm)
            c.setFont("Helvetica-Bold", 7)
            c.drawString(24 * mm, 20 * mm, asset["asset_number"])
            c.setFont("Helvetica", 6)
            c.drawString(24 * mm, 16 * mm, asset["name"][:20])
        else:
            # Standard: Barcode + QR + Info
            # Barcode at top
            code_val = asset["asset_number"]
            bc_cls = barcode.get_barcode_class("code128")
            bc_obj = bc_cls(code_val, writer=ImageWriter())
            bc_buf = io.BytesIO()
            bc_obj.write(bc_buf, options={"write_text": False, "quiet_zone": 2, "module_height": 10})
            bc_buf.seek(0)
            c.drawImage(ImageReader(bc_buf), 5 * mm, LH - 22 * mm, width=80 * mm, height=18 * mm, preserveAspectRatio=True)
            
            # QR code at bottom right
            base_url = request.base_url.scheme + "://" + request.base_url.netloc
            qr_data = json.dumps({
                "type": "asset",
                "asset_id": asset["id"],
                "asset_number": asset["asset_number"],
                "name": asset["name"],
                "category": asset.get("category_name", ""),
                "location": asset.get("location", ""),
                "url": f"{base_url}/asset/{asset['id']}"
            })
            qr_obj = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=5, border=1)
            qr_obj.add_data(qr_data)
            qr_obj.make(fit=True)
            qr_img = qr_obj.make_image(fill_color="black", back_color="white")
            qr_buf = io.BytesIO()
            qr_img.save(qr_buf, format="PNG")
            qr_buf.seek(0)
            c.drawImage(ImageReader(qr_buf), LW - 25 * mm, 5 * mm, width=20 * mm, height=20 * mm)
            
            # Asset info
            c.setFont("Helvetica-Bold", 11)
            c.drawString(5 * mm, LH - 28 * mm, asset["name"][:40])
            c.setFont("Helvetica", 8)
            c.drawString(5 * mm, LH - 33 * mm, f"Kode: {asset['asset_number']}")
            c.drawString(5 * mm, LH - 37 * mm, f"Kategori: {asset.get('category_name', '-')}")
            c.drawString(5 * mm, LH - 41 * mm, f"Lokasi: {asset.get('location', '-')}")
            c.setFont("Helvetica", 6)
            c.drawString(5 * mm, 3 * mm, f"CV. Dewi Aditya • {asset.get('purchase_date', '')}")
        
        c.showPage()
        c.save()
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="asset-label-{asset["asset_number"]}.pdf"'}
        )
    except Exception as e:
        logger.error(f"Label PDF generation error: {e}")
        raise HTTPException(500, "Gagal generate label PDF")


# ═══════════════════════════════════════════════════════════════════════════════
# UTILIZATION REPORT (Session 28 — additional)
# Computes how effectively assets are being used (assignment coverage,
# idle time, top-utilized categories/assets/assignees, underutilized list).
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_date_yyyymmdd(s: str, default: date) -> date:
    if not s:
        return default
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        return default


def _days_between(d1: date, d2: date) -> int:
    """Inclusive day count between d1 and d2 (d2 >= d1)."""
    if d2 < d1:
        return 0
    return (d2 - d1).days + 1


def _intersect_days(a_start: date, a_end: date, b_start: date, b_end: date) -> int:
    """Number of overlap days between [a_start..a_end] and [b_start..b_end]."""
    start = max(a_start, b_start)
    end = min(a_end, b_end)
    if end < start:
        return 0
    return (end - start).days + 1


@router.get("/reports/utilization")
async def asset_utilization_report(
    request: Request,
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD (default: 90 days ago)"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD (default: today)"),
    category_id: Optional[str] = Query(None),
    assignee_id: Optional[str] = Query(None),
    underutilized_threshold: int = Query(30, ge=0, le=100, description="Asset is underutilized if utilization% < this"),
    limit: int = Query(100, ge=1, le=500, description="Max assets returned in detailed list"),
):
    """
    Asset Utilization Report.

    Computes per-asset utilization% within [date_from..date_to] window:
        utilization% = (sum of assigned days within window / window days) × 100

    Returns:
        - summary       : aggregate KPIs
        - by_category   : utilization broken down per category
        - by_assignee   : top users holding most assets / days
        - top_utilized  : assets with highest utilization
        - underutilized : assets below threshold
        - idle_assets   : assets with zero assignment in window
        - status_breakdown: current status counts (active/in_maint/disposed)
    """
    await require_auth(request)
    db = get_db()

    today = date.today()
    win_start = _parse_date_yyyymmdd(date_from, today - relativedelta(days=90))
    win_end = _parse_date_yyyymmdd(date_to, today)
    if win_end < win_start:
        raise HTTPException(400, "date_to harus >= date_from")
    win_days = _days_between(win_start, win_end)

    # ─── Fetch assets ────────────────────────────────────────────────────────
    asset_filter: dict = {"status": {"$ne": "disposed"}}
    if category_id:
        asset_filter["category_id"] = category_id
    assets = await db.dewi_assets.find(
        asset_filter,
        {
            "_id": 0,
            "id": 1, "name": 1, "asset_number": 1, "category_id": 1, "category_name": 1,
            "status": 1, "purchase_cost": 1, "purchase_date": 1,
            "assigned_to_id": 1, "assigned_to_name": 1,
        }
    ).to_list(5000)

    asset_ids = [a["id"] for a in assets]
    if not asset_ids:
        return {
            "window": {"date_from": win_start.isoformat(), "date_to": win_end.isoformat(), "days": win_days},
            "summary": {
                "total_assets": 0, "assets_in_use_today": 0,
                "avg_utilization_pct": 0, "fully_utilized_count": 0,
                "underutilized_count": 0, "idle_in_window_count": 0,
                "total_purchase_cost": 0, "underutilized_value_at_risk": 0,
            },
            "by_category": [], "by_assignee": [],
            "top_utilized": [], "underutilized": [], "idle_assets": [],
            "status_breakdown": {"active": 0, "in_maintenance": 0, "disposed": 0},
        }

    # ─── Fetch assignment history (intersecting the window) ──────────────────
    assn_filter: dict = {
        "asset_id": {"$in": asset_ids},
        # we want assignments that are either still open OR ended after window start
        "$or": [
            {"returned_date": None},
            {"returned_date": {"$gte": win_start.isoformat()}},
        ],
        # AND started before window end (use assigned_date — string YYYY-MM-DD)
        "assigned_date": {"$lte": win_end.isoformat()},
    }
    if assignee_id:
        assn_filter["assigned_to_id"] = assignee_id

    assignments = await db.dewi_asset_assignments.find(
        assn_filter, {"_id": 0}
    ).to_list(20000)

    # ─── Compute per-asset utilization ───────────────────────────────────────
    asset_util: dict = {}
    for a in assets:
        asset_util[a["id"]] = {
            "asset": a,
            "assigned_days": 0,
            "assignment_count": 0,
            "current_assignee": a.get("assigned_to_name") or None,
            "current_assignee_id": a.get("assigned_to_id") or None,
            "last_assigned_date": None,
            "last_returned_date": None,
        }

    for ass in assignments:
        aid = ass.get("asset_id")
        if aid not in asset_util:
            continue
        a_start = _parse_date_yyyymmdd(ass.get("assigned_date"), win_start)
        a_end_str = ass.get("returned_date")
        a_end = _parse_date_yyyymmdd(a_end_str, today) if a_end_str else today
        # cap to window
        overlap = _intersect_days(a_start, a_end, win_start, win_end)
        bucket = asset_util[aid]
        bucket["assigned_days"] += overlap
        bucket["assignment_count"] += 1
        if not bucket["last_assigned_date"] or ass.get("assigned_date", "") > bucket["last_assigned_date"]:
            bucket["last_assigned_date"] = ass.get("assigned_date")
        if a_end_str and (not bucket["last_returned_date"] or a_end_str > bucket["last_returned_date"]):
            bucket["last_returned_date"] = a_end_str

    # Compute % and classify
    enriched = []
    full_count = 0
    under_count = 0
    idle_count = 0
    total_util_sum = 0.0
    in_use_today = 0
    total_value_at_risk = 0.0

    for aid, b in asset_util.items():
        # Window can clip by purchase_date if asset purchased mid-window
        purchase_date = _parse_date_yyyymmdd(b["asset"].get("purchase_date"), win_start)
        effective_start = max(win_start, purchase_date)
        effective_days = _days_between(effective_start, win_end)
        if effective_days <= 0:
            util_pct = 0.0
        else:
            # Cap assigned_days to effective_days (assignment cannot be > available days)
            asgn_days = min(b["assigned_days"], effective_days)
            util_pct = round(asgn_days / effective_days * 100, 1)

        b["utilization_pct"] = util_pct
        b["effective_window_days"] = effective_days
        b["asset_id"] = aid
        b["asset_name"] = b["asset"].get("name")
        b["asset_number"] = b["asset"].get("asset_number")
        b["category_id"] = b["asset"].get("category_id")
        b["category_name"] = b["asset"].get("category_name")
        b["purchase_cost"] = b["asset"].get("purchase_cost") or 0
        b["status"] = b["asset"].get("status")

        total_util_sum += util_pct
        if util_pct >= 95:
            full_count += 1
        if util_pct < underutilized_threshold:
            under_count += 1
            total_value_at_risk += b["purchase_cost"]
        if b["assigned_days"] == 0:
            idle_count += 1
        if b["current_assignee_id"]:
            in_use_today += 1

        # Drop nested asset key to keep payload light
        b.pop("asset", None)
        enriched.append(b)

    total_assets = len(enriched)
    avg_util = round(total_util_sum / total_assets, 1) if total_assets else 0.0

    # ─── Sort detailed lists ─────────────────────────────────────────────────
    enriched_sorted_desc = sorted(enriched, key=lambda x: x["utilization_pct"], reverse=True)
    enriched_sorted_asc = sorted(enriched, key=lambda x: x["utilization_pct"])

    top_utilized = enriched_sorted_desc[:limit]
    underutilized = [e for e in enriched_sorted_asc if e["utilization_pct"] < underutilized_threshold][:limit]
    idle_assets = [e for e in enriched if e["assigned_days"] == 0][:limit]

    # ─── By category aggregation ─────────────────────────────────────────────
    cat_map: dict = {}
    for e in enriched:
        cid = e["category_id"] or "_uncategorized_"
        cname = e["category_name"] or "Uncategorized"
        if cid not in cat_map:
            cat_map[cid] = {
                "category_id": cid, "category_name": cname,
                "asset_count": 0, "total_utilization_sum": 0.0,
                "underutilized_count": 0, "idle_count": 0,
                "total_purchase_cost": 0.0,
            }
        c = cat_map[cid]
        c["asset_count"] += 1
        c["total_utilization_sum"] += e["utilization_pct"]
        c["total_purchase_cost"] += e["purchase_cost"]
        if e["utilization_pct"] < underutilized_threshold:
            c["underutilized_count"] += 1
        if e["assigned_days"] == 0:
            c["idle_count"] += 1

    by_category = []
    for c in cat_map.values():
        avg = round(c["total_utilization_sum"] / c["asset_count"], 1) if c["asset_count"] else 0
        by_category.append({
            "category_id": c["category_id"],
            "category_name": c["category_name"],
            "asset_count": c["asset_count"],
            "avg_utilization_pct": avg,
            "underutilized_count": c["underutilized_count"],
            "idle_count": c["idle_count"],
            "total_purchase_cost": round(c["total_purchase_cost"], 2),
        })
    by_category.sort(key=lambda x: x["avg_utilization_pct"], reverse=True)

    # ─── By assignee aggregation ─────────────────────────────────────────────
    assignee_map: dict = {}
    for ass in assignments:
        aid = ass.get("assigned_to_id") or "_unknown_"
        aname = ass.get("assigned_to_name") or "—"
        if aid not in assignee_map:
            assignee_map[aid] = {
                "assignee_id": aid, "assignee_name": aname,
                "asset_count": set(), "total_assigned_days": 0,
            }
        a_start = _parse_date_yyyymmdd(ass.get("assigned_date"), win_start)
        a_end = _parse_date_yyyymmdd(ass.get("returned_date"), today) if ass.get("returned_date") else today
        overlap = _intersect_days(a_start, a_end, win_start, win_end)
        assignee_map[aid]["asset_count"].add(ass.get("asset_id"))
        assignee_map[aid]["total_assigned_days"] += overlap

    by_assignee = []
    for a in assignee_map.values():
        by_assignee.append({
            "assignee_id": a["assignee_id"],
            "assignee_name": a["assignee_name"],
            "unique_assets": len(a["asset_count"]),
            "total_assigned_days": a["total_assigned_days"],
        })
    by_assignee.sort(key=lambda x: x["total_assigned_days"], reverse=True)
    by_assignee = by_assignee[:50]

    # ─── Status breakdown ────────────────────────────────────────────────────
    status_breakdown = {
        "active": sum(1 for e in enriched if e["status"] == "active"),
        "in_maintenance": sum(1 for e in enriched if e["status"] == "in_maintenance"),
        "disposed": 0,  # disposed already excluded
    }

    total_purchase_cost = round(sum(e["purchase_cost"] for e in enriched), 2)

    return {
        "window": {
            "date_from": win_start.isoformat(),
            "date_to": win_end.isoformat(),
            "days": win_days,
        },
        "filters": {
            "category_id": category_id,
            "assignee_id": assignee_id,
            "underutilized_threshold": underutilized_threshold,
        },
        "summary": {
            "total_assets": total_assets,
            "assets_in_use_today": in_use_today,
            "avg_utilization_pct": avg_util,
            "fully_utilized_count": full_count,
            "underutilized_count": under_count,
            "idle_in_window_count": idle_count,
            "total_purchase_cost": total_purchase_cost,
            "underutilized_value_at_risk": round(total_value_at_risk, 2),
        },
        "by_category": by_category,
        "by_assignee": by_assignee,
        "top_utilized": top_utilized,
        "underutilized": underutilized,
        "idle_assets": idle_assets,
        "status_breakdown": status_breakdown,
    }


@router.get("/reports/utilization/export.csv")
async def asset_utilization_export_csv(
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    category_id: Optional[str] = Query(None),
    assignee_id: Optional[str] = Query(None),
    underutilized_threshold: int = Query(30, ge=0, le=100),
):
    """Export full utilization report as CSV."""
    from fastapi.responses import Response
    import csv
    import io

    rpt = await asset_utilization_report(
        request,
        date_from=date_from,
        date_to=date_to,
        category_id=category_id,
        assignee_id=assignee_id,
        underutilized_threshold=underutilized_threshold,
        limit=5000,
    )

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([
        "Asset Number", "Asset Name", "Category", "Status",
        "Utilization %", "Assigned Days", "Effective Days",
        "Assignment Count", "Current Assignee",
        "Last Assigned Date", "Last Returned Date",
        "Purchase Cost",
    ])
    for e in rpt["top_utilized"]:
        w.writerow([
            e.get("asset_number") or "",
            e.get("asset_name") or "",
            e.get("category_name") or "",
            e.get("status") or "",
            e.get("utilization_pct"),
            e.get("assigned_days"),
            e.get("effective_window_days"),
            e.get("assignment_count"),
            e.get("current_assignee") or "",
            e.get("last_assigned_date") or "",
            e.get("last_returned_date") or "",
            e.get("purchase_cost"),
        ])
    out.seek(0)

    filename = f"asset_utilization_{rpt['window']['date_from']}_to_{rpt['window']['date_to']}.csv"
    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ─── Init indexes ─────────────────────────────────────────────────────────

async def create_asset_indexes(db):
    await db.dewi_assets.create_index([("status", 1), ("category_id", 1)])
    await db.dewi_assets.create_index([("asset_number", 1)], unique=True, sparse=True)
    await db.dewi_assets.create_index([("assigned_to_id", 1)])
    await db.dewi_asset_depreciation.create_index([("asset_id", 1), ("period", 1)], unique=True)
    await db.dewi_asset_assignments.create_index([("asset_id", 1), ("status", 1)])
    await db.dewi_asset_maintenance.create_index([("asset_id", 1), ("maintenance_date", -1)])
    # Predictive maintenance acknowledgment
    await db.dewi_asset_pm_acknowledgments.create_index([("asset_id", 1), ("acknowledged_at", -1)])


# ═══════════════════════════════════════════════════════════════════════════════
# PREDICTIVE MAINTENANCE ALERT (Session 28)
# Heuristic-based maintenance forecasting:
#   - Overdue: next_scheduled date passed
#   - Upcoming: next_scheduled within 30 days
#   - Stale: no maintenance for > stale_months
#   - High frequency: > N maintenance in last 90 days (anomaly)
#   - Predicted: computed avg-interval predicted_next_due_date
# ═══════════════════════════════════════════════════════════════════════════════

def _safe_avg_interval_days(dates_iso_desc):
    """
    Given a list of maintenance date strings (YYYY-MM-DD),
    return avg interval in days between consecutive maintenances.
    Returns None if < 2 dates.
    """
    if len(dates_iso_desc) < 2:
        return None
    parsed = []
    for d in dates_iso_desc:
        try:
            parsed.append(datetime.strptime(str(d)[:10], "%Y-%m-%d").date())
        except Exception:
            continue
    if len(parsed) < 2:
        return None
    parsed.sort(reverse=True)
    diffs = []
    for i in range(len(parsed) - 1):
        diffs.append((parsed[i] - parsed[i + 1]).days)
    diffs = [d for d in diffs if d > 0]
    if not diffs:
        return None
    return sum(diffs) / len(diffs)


@router.get("/predictive-maintenance/alerts")
async def predictive_maintenance_alerts(
    request: Request,
    upcoming_window_days: int = Query(30, ge=1, le=365),
    stale_months: int = Query(6, ge=1, le=36),
    high_frequency_window_days: int = Query(90, ge=7, le=365),
    high_frequency_threshold: int = Query(3, ge=2, le=20),
    category_id: Optional[str] = Query(None),
):
    """
    Predictive Maintenance Alert engine.
    Returns categorized alerts: overdue, upcoming, stale, high_frequency, predicted.
    """
    await require_auth(request)
    db = get_db()
    today = date.today()
    today_iso = today.isoformat()

    asset_filter = {"status": {"$ne": "disposed"}}
    if category_id:
        asset_filter["category_id"] = category_id

    assets = await db.dewi_assets.find(
        asset_filter,
        {
            "_id": 0,
            "id": 1, "name": 1, "asset_number": 1, "category_id": 1, "category_name": 1,
            "status": 1, "purchase_cost": 1, "purchase_date": 1,
            "assigned_to_id": 1, "assigned_to_name": 1,
        }
    ).to_list(5000)

    if not assets:
        return {
            "generated_at": _now().isoformat(),
            "config": {
                "upcoming_window_days": upcoming_window_days,
                "stale_months": stale_months,
                "high_frequency_window_days": high_frequency_window_days,
                "high_frequency_threshold": high_frequency_threshold,
            },
            "summary": {
                "overdue_count": 0, "upcoming_count": 0, "stale_count": 0,
                "high_frequency_count": 0, "predicted_count": 0, "total_alerts": 0,
                "critical_count": 0,
            },
            "overdue": [], "upcoming": [], "stale": [],
            "high_frequency": [], "predicted": [],
        }

    asset_map = {a["id"]: a for a in assets}
    asset_ids = list(asset_map.keys())

    pipeline = [
        {"$match": {"asset_id": {"$in": asset_ids}}},
        {"$sort": {"maintenance_date": -1}},
        {"$group": {
            "_id": "$asset_id",
            "last_date": {"$first": "$maintenance_date"},
            "last_status": {"$first": "$status"},
            "last_type": {"$first": "$type"},
            "last_next_scheduled": {"$first": "$next_scheduled"},
            "last_cost": {"$first": "$cost"},
            "history": {"$push": {
                "date": "$maintenance_date",
                "cost": "$cost",
                "type": "$type",
                "next_scheduled": "$next_scheduled",
            }},
            "total_count": {"$sum": 1},
        }}
    ]
    maint_agg = await db.dewi_asset_maintenance.aggregate(pipeline).to_list(5000)
    maint_map = {m["_id"]: m for m in maint_agg}

    hf_window_start = (today - relativedelta(days=high_frequency_window_days)).isoformat()
    recent_pipeline = [
        {"$match": {
            "asset_id": {"$in": asset_ids},
            "maintenance_date": {"$gte": hf_window_start},
        }},
        {"$group": {"_id": "$asset_id", "recent_count": {"$sum": 1},
                    "recent_total_cost": {"$sum": "$cost"}}}
    ]
    recent_agg = await db.dewi_asset_maintenance.aggregate(recent_pipeline).to_list(5000)
    recent_map = {r["_id"]: r for r in recent_agg}

    ack_recent = await db.dewi_asset_pm_acknowledgments.find(
        {"acknowledged_at": {"$gte": (_now() - relativedelta(days=30)).isoformat()}},
        {"_id": 0, "asset_id": 1, "alert_kind": 1}
    ).to_list(5000)
    ack_set = {(a["asset_id"], a["alert_kind"]) for a in ack_recent}

    overdue, upcoming, stale, high_frequency, predicted = [], [], [], [], []

    stale_cutoff = (today - relativedelta(months=stale_months)).isoformat()
    upcoming_cutoff = (today + relativedelta(days=upcoming_window_days)).isoformat()

    for aid, asset in asset_map.items():
        m = maint_map.get(aid)
        recent = recent_map.get(aid)
        base = {
            "asset_id": aid,
            "asset_name": asset.get("name"),
            "asset_number": asset.get("asset_number"),
            "category_id": asset.get("category_id"),
            "category_name": asset.get("category_name"),
            "current_assignee": asset.get("assigned_to_name"),
            "current_status": asset.get("status"),
            "purchase_cost": asset.get("purchase_cost") or 0,
        }

        if m and m.get("last_next_scheduled") and str(m["last_next_scheduled"])[:10] < today_iso:
            days_overdue = (today - datetime.strptime(str(m["last_next_scheduled"])[:10], "%Y-%m-%d").date()).days
            if (aid, "overdue") not in ack_set:
                overdue.append({
                    **base,
                    "kind": "overdue",
                    "severity": "critical" if days_overdue > 14 else "warning",
                    "scheduled_date": str(m["last_next_scheduled"])[:10],
                    "days_overdue": days_overdue,
                    "last_maintenance_date": m["last_date"],
                    "last_type": m["last_type"],
                    "recommended_action": "Lakukan maintenance segera. Aset telah melewati jadwal.",
                })
        elif m and m.get("last_next_scheduled") and today_iso <= str(m["last_next_scheduled"])[:10] <= upcoming_cutoff:
            days_until = (datetime.strptime(str(m["last_next_scheduled"])[:10], "%Y-%m-%d").date() - today).days
            if (aid, "upcoming") not in ack_set:
                upcoming.append({
                    **base,
                    "kind": "upcoming",
                    "severity": "warning" if days_until <= 7 else "info",
                    "scheduled_date": str(m["last_next_scheduled"])[:10],
                    "days_until": days_until,
                    "last_maintenance_date": m["last_date"],
                    "last_type": m["last_type"],
                    "recommended_action": f"Siapkan maintenance dalam {days_until} hari.",
                })

        if m and m.get("last_date") and m["last_date"] < stale_cutoff:
            already_alerted = any(o["asset_id"] == aid for o in overdue) or any(u["asset_id"] == aid for u in upcoming)
            if not already_alerted and (aid, "stale") not in ack_set:
                last_date = datetime.strptime(str(m["last_date"])[:10], "%Y-%m-%d").date()
                days_since = (today - last_date).days
                months_since = round(days_since / 30.5, 1)
                stale.append({
                    **base,
                    "kind": "stale",
                    "severity": "warning" if months_since > stale_months * 1.5 else "info",
                    "last_maintenance_date": m["last_date"],
                    "days_since_maintenance": days_since,
                    "months_since_maintenance": months_since,
                    "last_type": m["last_type"],
                    "total_maintenance_count": m["total_count"],
                    "recommended_action": f"Jadwalkan inspeksi rutin. Sudah {months_since} bulan tanpa maintenance.",
                })

        if recent and recent["recent_count"] >= high_frequency_threshold:
            if (aid, "high_frequency") not in ack_set:
                high_frequency.append({
                    **base,
                    "kind": "high_frequency",
                    "severity": "critical" if recent["recent_count"] >= high_frequency_threshold * 2 else "warning",
                    "recent_count": recent["recent_count"],
                    "recent_total_cost": round(recent["recent_total_cost"] or 0, 2),
                    "window_days": high_frequency_window_days,
                    "last_maintenance_date": m["last_date"] if m else None,
                    "recommended_action": (
                        f"Pola maintenance abnormal terdeteksi: {recent['recent_count']}x dalam "
                        f"{high_frequency_window_days} hari. Pertimbangkan replacement atau "
                        f"investigasi root cause."
                    ),
                })

        if m and not m.get("last_next_scheduled") and m.get("total_count", 0) >= 2:
            dates = [h["date"] for h in m["history"] if h.get("date")]
            avg_days = _safe_avg_interval_days(dates)
            if avg_days and avg_days > 0:
                last_date = datetime.strptime(str(m["last_date"])[:10], "%Y-%m-%d").date()
                predicted_due = last_date + relativedelta(days=int(avg_days))
                if predicted_due <= (today + relativedelta(days=upcoming_window_days)):
                    already_alerted = any(o["asset_id"] == aid for o in overdue) or any(u["asset_id"] == aid for u in upcoming)
                    if not already_alerted and (aid, "predicted") not in ack_set:
                        days_offset = (predicted_due - today).days
                        predicted.append({
                            **base,
                            "kind": "predicted",
                            "severity": "warning" if days_offset < 0 else "info",
                            "last_maintenance_date": m["last_date"],
                            "avg_interval_days": round(avg_days, 1),
                            "predicted_next_due_date": predicted_due.isoformat(),
                            "days_offset": days_offset,
                            "total_maintenance_count": m["total_count"],
                            "recommended_action": (
                                f"Berdasarkan pola historical ({m['total_count']}x maintenance, "
                                f"avg {round(avg_days)} hari interval), maintenance berikutnya "
                                f"diprediksi {'sudah lewat' if days_offset < 0 else f'dalam {days_offset} hari'}."
                            ),
                        })

    severity_order = {"critical": 0, "warning": 1, "info": 2}
    overdue.sort(key=lambda x: (-x["days_overdue"], severity_order.get(x["severity"], 9)))
    upcoming.sort(key=lambda x: (x["days_until"], severity_order.get(x["severity"], 9)))
    stale.sort(key=lambda x: (-x["days_since_maintenance"], severity_order.get(x["severity"], 9)))
    high_frequency.sort(key=lambda x: (-x["recent_count"], severity_order.get(x["severity"], 9)))
    predicted.sort(key=lambda x: (x["days_offset"], severity_order.get(x["severity"], 9)))

    summary = {
        "overdue_count": len(overdue),
        "upcoming_count": len(upcoming),
        "stale_count": len(stale),
        "high_frequency_count": len(high_frequency),
        "predicted_count": len(predicted),
        "total_alerts": len(overdue) + len(upcoming) + len(stale) + len(high_frequency) + len(predicted),
        "critical_count": sum(1 for li in [overdue, upcoming, stale, high_frequency, predicted]
                               for x in li if x.get("severity") == "critical"),
    }

    return {
        "generated_at": _now().isoformat(),
        "config": {
            "upcoming_window_days": upcoming_window_days,
            "stale_months": stale_months,
            "high_frequency_window_days": high_frequency_window_days,
            "high_frequency_threshold": high_frequency_threshold,
        },
        "summary": summary,
        "overdue": overdue,
        "upcoming": upcoming,
        "stale": stale,
        "high_frequency": high_frequency,
        "predicted": predicted,
    }


@router.post("/predictive-maintenance/acknowledge")
async def acknowledge_pm_alert(request: Request):
    """
    Acknowledge a predictive maintenance alert so it stops appearing for 30 days.
    Body: { asset_id, alert_kind, note? }
    """
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    asset_id = body.get("asset_id")
    alert_kind = body.get("alert_kind")
    if not asset_id or not alert_kind:
        raise HTTPException(400, "asset_id dan alert_kind wajib diisi")
    if alert_kind not in {"overdue", "upcoming", "stale", "high_frequency", "predicted"}:
        raise HTTPException(400, "alert_kind tidak valid")

    doc = {
        "id": _uid(),
        "asset_id": asset_id,
        "alert_kind": alert_kind,
        "note": (body.get("note") or "").strip(),
        "acknowledged_by_id": user["id"],
        "acknowledged_by_name": user.get("name", ""),
        "acknowledged_at": _now().isoformat(),
    }
    await db.dewi_asset_pm_acknowledgments.insert_one(doc)
    return _ser(doc)


@router.get("/predictive-maintenance/acknowledgments")
async def list_pm_acknowledgments(
    request: Request,
    asset_id: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
):
    """List recent PM acknowledgments."""
    await require_auth(request)
    db = get_db()
    flt: dict = {}
    if asset_id:
        flt["asset_id"] = asset_id
    rows = await db.dewi_asset_pm_acknowledgments.find(
        flt, {"_id": 0}
    ).sort("acknowledged_at", -1).limit(limit).to_list(limit)
    return [_ser(r) for r in rows]

